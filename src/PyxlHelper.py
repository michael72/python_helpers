'''
Created on 16.02.2020

@author: Michael Schulte
'''
from operator import getitem
import os
import sys
from time import sleep
import traceback
import zipfile

from FileHelper import makeWritable

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl import load_workbook
    pyxl_enabled = True
except ImportError:
    pyxl_enabled = False
    sys.stderr.write("Failed to import openpyxl library - install it via : pip install openpyxl\npip is located in subfolder Scripts of Python-Installation")

class PyxlWorkbook(object):
    '''
    Helper class for workbooks.
    
    Use:
    with PyxlWorkbook(filename) as wb:
        # do the changes on wb

    after with-statement the workbook is automatically saved to the filename with 
    auto adjusted column width.
            
    Alternative use:
    wb = PyxlWorkbook(filename)
    wb.prepare()
    # ... work on wb
    wb.close()
        
    Also some helper functions are provided.
    
    '''
    def __init__(self, filename, read_only=False, update=False):
        ''' Creates the helper which can be used as workbook.
        :filename: full path to excel file
        :read_only: open the existing file in read-only mode. http: files are automatically downloaded and saved as temporary file.
        :update: create or update the file - if it exists. Default = False: overwrite the entire file (e.g. when generating a report).
        '''
        self.filename = filename
        self.fh = None
        self.wb = Workbook()
        self.readOnly = read_only
        self.url = ''
        assert not (read_only and update), "can only set read_only or update - not both of them!"
        self.update = update and os.path.exists(filename)

    def __getattr__(self, attr):
        # delegate all (other) attributes to the actual workbook
        return getattr(self.wb, attr)
    
    def __getitem__(self, *args):
        ''' delegate all access with [] to the workbook '''
        return getitem(self.wb, *args)

    def __enter__(self):
        if self.fh == None:
            self.prepare()
        return self

    def __exit__(self, exc_type, value, tb):
        if value == None:
            _,value,tb = sys.exc_info()
        if value == None:
            self.close()
            return True
        if exc_type == None:
            exc_type = BaseException("unknown exception type")
        traceback.print_exception(exc_type, value, tb)                
                    
    def _trySave(self):
        printError = True
        while True:
            try:
                self.wb.save(self.filename)
                break
            except IOError as ex:
                if printError:
                    if not os.path.exists(os.path.dirname(self.filename)):
                        sys.stderr.write("Invalid path %s\n" % (self.filename))
                        raise(ex)
                    sys.stderr.write("File access error: Please close the excel document first!!\n")
                    printError = False
                sleep(0.1)
    
    def _tryOpen(self):
        printError = True
        assert(not self.readOnly)
        makeWritable(self.filename)

        while True:
            try:
                self.fh = open(self.filename, 'a' if self.update else 'w') # ensure write access during processing
                if not printError:
                    sys.stderr.write("... closed\n")
                break
            except IOError as ex:
                if os.path.exists(os.path.dirname(self.filename)):
                    if printError:
                        sys.stderr.write("File access error: Please close the excel document first!!\n")
                        printError = False
                    sleep(0.1)
                else:
                    raise ex
        if self.update:
            try:
                self.wb = load_workbook(self.filename, keep_vba=self.filename.endswith('.xlsm'))
            except zipfile.BadZipfile:
                sys.stderr.write("Could not open Excel-file - recreating it!\n")
                self.update = False  
    
    def close(self):
        if self.readOnly:
            # hack for bug in openpyxl
            try:
                self.wb._archive.close()
            except:
                pass

        else:
            for ws in self.worksheets:
                self.adjustColumns(ws)
            if self.fh != None:
                try:
                    os.close(self.fh)
                    self.fh = None
                except:
                    pass
            self._trySave()
        
    def prepare(self):
        self.wb = Workbook()
        exists = os.path.exists(self.filename)
        if exists:
            self._tryOpen()

        return self # for chained calls  
    
    def createUrl(self, url, label):                        
        return '=HYPERLINK("{}", "{}")'.format(url, label)

    def mkBold(self, ws, row):
        assert(not self.readOnly)
        rows = list(ws.rows)
        if row <= len(rows):
            for cell in rows[row-1]:
                cell.font = Font(bold=True)
                
    def mkSmall(self, cell, sz=4):
        assert(not self.readOnly)
        cell.font = Font(sz=sz)

    @staticmethod
    def content(cell):
        ''' get the content of the cell without formatting / URL '''
        try:
            s = str(cell.value)
            if cell.style == 'Output' or (cell.font and ((not cell.font.sz) or (cell.font.sz < 8))):
                return 0 
            if s.startswith('=') and '"' in s: # remove Hyperlink and other formula stuff
                idx = s.rfind('"')
                s = s[s.rfind('"', 0, idx-1)+1:idx]
            elif cell.style == 'Hyperlink':
                return ""
        except UnicodeEncodeError:
            return ""
        return s
                    
    @classmethod
    def cellLen(cls, cell):
        ''' length of the cell content (to calculate the needed column width)'''
        return len(cls.content(cell))
            
    def adjustColumns(self, ws=None):
        ''' tries to adjust the cell width to fit the content '''
        assert(not self.readOnly)
        if ws == None:
            ws = self.active
        for col in ws.columns:
            max_length = 0
            for cell in col:
                if self.cellLen(cell) > max_length:
                    max_length = self.cellLen(cell)
            adjusted_width = (max_length + 1) * 1.4 # this is PI * thumb
            ws.column_dimensions[col[0].column].width = adjusted_width
        return self

